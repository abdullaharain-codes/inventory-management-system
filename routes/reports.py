import csv
import io
from datetime import datetime
from flask import Blueprint, request, jsonify, Response
from db.connection import get_db_connection
from mysql.connector import Error
from middleware.auth_middleware import roles_required
from utils.report_pdf_generator import generate_report_pdf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

reports_bp = Blueprint('reports', __name__)

VALID_GRANULARITIES = ('daily', 'weekly', 'monthly', 'yearly')


def _parse_dates(args):
    """Extract and validate date_from, date_to from request args."""
    date_from = args.get('date_from', '').strip()
    date_to = args.get('date_to', '').strip()
    if not date_from or not date_to:
        return None, None, 'date_from and date_to are required'
    try:
        datetime.strptime(date_from, '%Y-%m-%d')
        datetime.strptime(date_to, '%Y-%m-%d')
    except ValueError:
        return None, None, 'Invalid date format. Use YYYY-MM-DD'
    if date_from > date_to:
        return None, None, 'date_from must be before or equal to date_to'
    return date_from, date_to, None


def _get_conn():
    conn = get_db_connection()
    if not conn:
        return None
    return conn


def _get_stock_report_data(cur, date_from, date_to, product_id_filter=None):
    """Shared logic for stock report queries.

    Returns (period_summary, by_product, movement_breakdown).
    ``cur`` must be a dictionary cursor.
    """
    if product_id_filter:
        prod_filter_sql = 'AND sl.product_id = %s'
        prod_filter_params = (product_id_filter,)
    else:
        prod_filter_sql = ''
        prod_filter_params = ()

    # Opening stock: latest entry BEFORE date_from
    cur.execute(f"""
        SELECT sl.product_id,
               p.name AS product_name,
               sl.quantity_before AS opening_stock
        FROM stock_ledger sl
        JOIN products p ON p.product_id = sl.product_id
        INNER JOIN (
            SELECT product_id, MAX(ledger_id) AS max_ledger
            FROM stock_ledger
            WHERE DATE(created_at) < %s
            GROUP BY product_id
        ) latest ON sl.ledger_id = latest.max_ledger
        WHERE 1=1 {prod_filter_sql}
    """, (date_from,) + prod_filter_params)
    opening_rows = {r['product_id']: r for r in cur.fetchall()}

    # Fallback: earliest-ever entry for products with no pre-date_from history
    cur.execute(f"""
        SELECT sl.product_id,
               sl.quantity_before AS opening_stock
        FROM stock_ledger sl
        INNER JOIN (
            SELECT product_id, MIN(ledger_id) AS min_ledger
            FROM stock_ledger
            GROUP BY product_id
        ) first ON sl.ledger_id = first.min_ledger
        WHERE sl.product_id NOT IN (
            SELECT product_id FROM stock_ledger WHERE DATE(created_at) < %s
        )
        {prod_filter_sql}
    """, (date_from,) + prod_filter_params)
    for r in cur.fetchall():
        if r['product_id'] not in opening_rows:
            opening_rows[r['product_id']] = {
                'product_id': r['product_id'],
                'product_name': '',
                'opening_stock': r['opening_stock'],
            }

    # Closing stock: latest entry at/before date_to
    cur.execute(f"""
        SELECT sl.product_id,
               sl.quantity_after AS closing_stock
        FROM stock_ledger sl
        INNER JOIN (
            SELECT product_id, MAX(ledger_id) AS max_ledger
            FROM stock_ledger
            WHERE DATE(created_at) <= %s
            GROUP BY product_id
        ) latest ON sl.ledger_id = latest.max_ledger
        WHERE 1=1 {prod_filter_sql}
    """, (date_to,) + prod_filter_params)
    closing_rows = {r['product_id']: r['closing_stock'] for r in cur.fetchall()}

    # In-period movements
    cur.execute(f"""
        SELECT sl.product_id,
               p.name AS product_name,
               COALESCE(SUM(CASE WHEN sl.quantity_change > 0 THEN sl.quantity_change ELSE 0 END), 0) AS stock_in,
               COALESCE(SUM(CASE WHEN sl.quantity_change < 0 THEN ABS(sl.quantity_change) ELSE 0 END), 0) AS stock_out
        FROM stock_ledger sl
        LEFT JOIN products p ON p.product_id = sl.product_id
        WHERE DATE(sl.created_at) BETWEEN %s AND %s
        {prod_filter_sql}
        GROUP BY sl.product_id, p.name
    """, (date_from, date_to) + prod_filter_params)
    movement_rows = cur.fetchall()

    # Build by_product
    all_product_ids = (
        set(opening_rows.keys())
        | set(closing_rows.keys())
        | set(r['product_id'] for r in movement_rows)
    )
    by_product = []
    total_in = 0
    total_out = 0
    for pid in sorted(all_product_ids):
        o = opening_rows.get(pid, {})
        c = closing_rows.get(pid, 0)
        m = next((r for r in movement_rows if r['product_id'] == pid), None)
        stock_in = m['stock_in'] if m else 0
        stock_out = m['stock_out'] if m else 0
        opening = o.get('opening_stock', 0) or 0
        name = o.get('product_name', '') or (m['product_name'] if m else '')
        expected_closing = opening + stock_in - stock_out
        by_product.append({
            'product_id': pid,
            'product_name': name,
            'opening_stock': opening,
            'stock_in': stock_in,
            'stock_out': stock_out,
            'closing_stock': c,
            'reconciled': (expected_closing == c),
        })
        total_in += stock_in
        total_out += stock_out

    # Movement breakdown
    cur.execute(f"""
        SELECT movement_type,
               COUNT(*) AS count,
               SUM(ABS(quantity_change)) AS total_quantity
        FROM stock_ledger sl
        WHERE DATE(sl.created_at) BETWEEN %s AND %s
        {prod_filter_sql}
        GROUP BY movement_type
        ORDER BY total_quantity DESC
    """, (date_from, date_to) + prod_filter_params)
    movement_breakdown = cur.fetchall()

    opening_total = sum(r['opening_stock'] for r in by_product)
    closing_total = sum(r['closing_stock'] for r in by_product)
    period_summary = {
        'total_stock_in': total_in,
        'total_stock_out': total_out,
        'net_change': total_in - total_out,
        'opening_stock_total': opening_total,
        'closing_stock_total': closing_total,
    }

    return period_summary, by_product, movement_breakdown


# ── Sales Report ──────────────────────────────────────────────────
@reports_bp.route('/api/reports/sales', methods=['GET'])
@roles_required('admin', 'manager')
def sales_report():
    date_from, date_to, err = _parse_dates(request.args)
    if err:
        return jsonify({'error': err}), 400

    granularity = request.args.get('granularity', 'daily').lower()
    if granularity not in VALID_GRANULARITIES:
        return jsonify({'error': f'granularity must be one of: {", ".join(VALID_GRANULARITIES)}'}), 400

    conn = _get_conn()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cur = conn.cursor(dictionary=True)

        # ── Period summary ──
        cur.execute("""
            SELECT COALESCE(SUM(grand_total), 0) AS total_revenue,
                   COUNT(*) AS total_bills,
                   COALESCE(SUM(discount_amount), 0) AS total_discount,
                   COALESCE(SUM(gst_amount), 0) AS total_gst,
                   COALESCE(SUM(subtotal), 0) AS total_subtotal
            FROM bills
            WHERE bill_date BETWEEN %s AND %s
        """, (date_from, date_to))
        ps = cur.fetchone()
        ps['avg_bill_value'] = round(ps['total_revenue'] / ps['total_bills'], 2) if ps['total_bills'] > 0 else 0
        ps['date_from'] = date_from
        ps['date_to'] = date_to

        # ── Time series ──
        if granularity == 'daily':
            group_expr = 'bill_date'
            label_expr = 'DATE_FORMAT(bill_date, "%Y-%m-%d")'
        elif granularity == 'weekly':
            group_expr = 'YEARWEEK(bill_date, 1)'
            label_expr = 'CONCAT(YEAR(bill_date), "-W", LPAD(WEEK(bill_date, 1), 2, "0"))'
        elif granularity == 'monthly':
            group_expr = 'DATE_FORMAT(bill_date, "%Y-%m")'
            label_expr = 'DATE_FORMAT(bill_date, "%Y-%m")'
        else:  # yearly
            group_expr = 'YEAR(bill_date)'
            label_expr = 'YEAR(bill_date)'

        cur.execute(f"""
            SELECT {label_expr} AS period_label,
                   MIN(bill_date) AS period_start,
                   MAX(bill_date) AS period_end,
                   COALESCE(SUM(grand_total), 0) AS revenue,
                   COUNT(*) AS bill_count,
                   COALESCE(SUM(discount_amount), 0) AS discount,
                   COALESCE(SUM(gst_amount), 0) AS gst
            FROM bills
            WHERE bill_date BETWEEN %s AND %s
            GROUP BY {group_expr}
            ORDER BY MIN(bill_date)
        """, (date_from, date_to))
        time_series = cur.fetchall()

        # ── Top 10 products by revenue ──
        cur.execute("""
            SELECT bi.product_id,
                   COALESCE(bi.product_name, p.name) AS product_name,
                   SUM(bi.quantity) AS quantity_sold,
                   SUM(bi.item_total) AS revenue
            FROM bill_items bi
            JOIN bills b ON b.bill_id = bi.bill_id
            LEFT JOIN products p ON p.product_id = bi.product_id
            WHERE b.bill_date BETWEEN %s AND %s
            GROUP BY bi.product_id, COALESCE(bi.product_name, p.name)
            ORDER BY revenue DESC
            LIMIT 10
        """, (date_from, date_to))
        top_products = cur.fetchall()

        # ── Payment method breakdown ──
        cur.execute("""
            SELECT payment_method,
                   COUNT(*) AS count,
                   COALESCE(SUM(grand_total), 0) AS revenue
            FROM bills
            WHERE bill_date BETWEEN %s AND %s
            GROUP BY payment_method
            ORDER BY revenue DESC
        """, (date_from, date_to))
        payment_breakdown = cur.fetchall()

        cur.close()
        return jsonify({
            'period_summary': ps,
            'time_series': time_series,
            'top_products': top_products,
            'payment_method_breakdown': payment_breakdown
        }), 200

    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


# ── Stock / Inventory Movement Report ─────────────────────────────
@reports_bp.route('/api/reports/stock', methods=['GET'])
@roles_required('admin', 'manager')
def stock_report():
    date_from, date_to, err = _parse_dates(request.args)
    if err:
        return jsonify({'error': err}), 400

    product_id_filter = request.args.get('product_id')
    if product_id_filter:
        try:
            product_id_filter = int(product_id_filter)
        except (ValueError, TypeError):
            return jsonify({'error': 'product_id must be an integer'}), 400

    category_id = request.args.get('category_id')
    if category_id:
        try:
            category_id = int(category_id)
        except (ValueError, TypeError):
            return jsonify({'error': 'category_id must be an integer'}), 400

    conn = _get_conn()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cur = conn.cursor(dictionary=True)

        # If category_id filter, first get matching product IDs
        product_ids_for_category = None
        if category_id:
            cur.execute("SELECT product_id FROM products WHERE category_id = %s", (category_id,))
            product_ids_for_category = [r['product_id'] for r in cur.fetchall()]
            if not product_ids_for_category:
                cur.close()
                return jsonify({
                    'period_summary': {'total_stock_in': 0, 'total_stock_out': 0, 'net_change': 0,
                                       'opening_stock_total': 0, 'closing_stock_total': 0},
                    'by_product': [], 'movement_breakdown': []
                }), 200

        period_summary, by_product, movement_breakdown = _get_stock_report_data(
            cur, date_from, date_to, product_id_filter
        )
        cur.close()

        # Apply category filter in-memory
        if product_ids_for_category is not None:
            by_product = [r for r in by_product if r['product_id'] in product_ids_for_category]
            # Recalculate summary
            total_in = sum(r['stock_in'] for r in by_product)
            total_out = sum(r['stock_out'] for r in by_product)
            opening_total = sum(r['opening_stock'] for r in by_product)
            closing_total = sum(r['closing_stock'] for r in by_product)
            period_summary = {
                'total_stock_in': total_in,
                'total_stock_out': total_out,
                'net_change': total_in - total_out,
                'opening_stock_total': opening_total,
                'closing_stock_total': closing_total,
            }

        return jsonify({
            'period_summary': period_summary,
            'by_product': by_product,
            'movement_breakdown': movement_breakdown
        }), 200

    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


# ── Profit Report ─────────────────────────────────────────────────
@reports_bp.route('/api/reports/profit', methods=['GET'])
@roles_required('admin', 'manager')
def profit_report():
    date_from, date_to, err = _parse_dates(request.args)
    if err:
        return jsonify({'error': err}), 400

    conn = _get_conn()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cur = conn.cursor(dictionary=True)

        # Per-product profit (only products with cost_price)
        cur.execute("""
            SELECT bi.product_id,
                   COALESCE(bi.product_name, p.name) AS product_name,
                   SUM(bi.quantity) AS quantity_sold,
                   SUM(bi.item_total) AS revenue,
                   SUM(bi.quantity * p.cost_price) AS cost,
                   SUM(bi.item_total - bi.quantity * p.cost_price) AS profit
            FROM bill_items bi
            JOIN bills b ON b.bill_id = bi.bill_id
            JOIN products p ON p.product_id = bi.product_id
            WHERE b.bill_date BETWEEN %s AND %s
              AND p.cost_price IS NOT NULL
            GROUP BY bi.product_id, COALESCE(bi.product_name, p.name)
            ORDER BY profit DESC
        """, (date_from, date_to))
        by_product = cur.fetchall()

        # Calculate margin_percent per product
        for row in by_product:
            if row['revenue'] and row['revenue'] > 0:
                row['margin_percent'] = round(float(row['profit']) / float(row['revenue']) * 100, 2)
            else:
                row['margin_percent'] = 0

        total_revenue = sum(float(r['revenue']) for r in by_product)
        total_cost = sum(float(r['cost']) for r in by_product)
        total_profit = sum(float(r['profit']) for r in by_product)
        avg_margin = round(total_profit / total_revenue * 100, 2) if total_revenue > 0 else 0

        # Count distinct products sold with NULL cost_price
        cur.execute("""
            SELECT COUNT(DISTINCT bi.product_id) AS excluded_count
            FROM bill_items bi
            JOIN bills b ON b.bill_id = bi.bill_id
            JOIN products p ON p.product_id = bi.product_id
            WHERE b.bill_date BETWEEN %s AND %s
              AND p.cost_price IS NULL
        """, (date_from, date_to))
        excluded = cur.fetchone()['excluded_count']

        period_summary = {
            'total_revenue': round(total_revenue, 2),
            'total_cost': round(total_cost, 2),
            'total_profit': round(total_profit, 2),
            'avg_margin_percent': avg_margin,
            'products_excluded_no_cost_price': excluded
        }

        cur.close()
        return jsonify({
            'period_summary': period_summary,
            'by_product': by_product
        }), 200

    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


# ── CSV: Sales Report ─────────────────────────────────────────────
@reports_bp.route('/api/reports/sales/csv', methods=['GET'])
@roles_required('admin', 'manager')
def sales_report_csv():
    date_from, date_to, err = _parse_dates(request.args)
    if err:
        return jsonify({'error': err}), 400

    granularity = request.args.get('granularity', 'daily').lower()
    if granularity not in VALID_GRANULARITIES:
        return jsonify({'error': f'granularity must be one of: {", ".join(VALID_GRANULARITIES)}'}), 400

    conn = _get_conn()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cur = conn.cursor(dictionary=True)

        if granularity == 'daily':
            group_expr = 'bill_date'
            label_expr = 'DATE_FORMAT(bill_date, "%Y-%m-%d")'
        elif granularity == 'weekly':
            group_expr = 'YEARWEEK(bill_date, 1)'
            label_expr = 'CONCAT(YEAR(bill_date), "-W", LPAD(WEEK(bill_date, 1), 2, "0"))'
        elif granularity == 'monthly':
            group_expr = 'DATE_FORMAT(bill_date, "%Y-%m")'
            label_expr = 'DATE_FORMAT(bill_date, "%Y-%m")'
        else:
            group_expr = 'YEAR(bill_date)'
            label_expr = 'YEAR(bill_date)'

        cur.execute(f"""
            SELECT {label_expr} AS period_label,
                   MIN(bill_date) AS period_start,
                   MAX(bill_date) AS period_end,
                   COALESCE(SUM(grand_total), 0) AS revenue,
                   COUNT(*) AS bill_count,
                   COALESCE(SUM(discount_amount), 0) AS discount,
                   COALESCE(SUM(gst_amount), 0) AS gst
            FROM bills
            WHERE bill_date BETWEEN %s AND %s
            GROUP BY {group_expr}
            ORDER BY MIN(bill_date)
        """, (date_from, date_to))
        rows = cur.fetchall()
        cur.close()

        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(['period_label', 'period_start', 'period_end', 'revenue', 'bill_count', 'discount', 'gst'])
        for r in rows:
            writer.writerow([r['period_label'], r['period_start'], r['period_end'],
                             r['revenue'], r['bill_count'], r['discount'], r['gst']])

        output = buf.getvalue()
        buf.close()
        filename = f"sales_report_{date_from}_to_{date_to}_{granularity}.csv"
        return Response(output, mimetype='text/csv',
                        headers={'Content-Disposition': f'attachment; filename="{filename}"'})
    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


# ── CSV: Stock Report ─────────────────────────────────────────────
@reports_bp.route('/api/reports/stock/csv', methods=['GET'])
@roles_required('admin', 'manager')
def stock_report_csv():
    date_from, date_to, err = _parse_dates(request.args)
    if err:
        return jsonify({'error': err}), 400

    category_id = request.args.get('category_id')
    if category_id:
        try:
            category_id = int(category_id)
        except (ValueError, TypeError):
            return jsonify({'error': 'category_id must be an integer'}), 400

    conn = _get_conn()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cur = conn.cursor(dictionary=True)
        period_summary, by_product, movement_breakdown = _get_stock_report_data(
            cur, date_from, date_to
        )
        cur.close()

        if category_id:
            pids = set()
            conn2 = _get_conn()
            if conn2:
                cur2 = conn2.cursor(dictionary=True)
                cur2.execute("SELECT product_id FROM products WHERE category_id = %s", (category_id,))
                pids = {r['product_id'] for r in cur2.fetchall()}
                cur2.close()
                conn2.close()
            by_product = [r for r in by_product if r['product_id'] in pids]

        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(['product_id', 'product_name', 'opening_stock', 'stock_in', 'stock_out', 'closing_stock', 'reconciled'])
        for row in by_product:
            writer.writerow([
                row['product_id'], row['product_name'], row['opening_stock'],
                row['stock_in'], row['stock_out'], row['closing_stock'], row['reconciled']
            ])

        output = buf.getvalue()
        buf.close()
        filename = f"stock_report_{date_from}_to_{date_to}.csv"
        return Response(output, mimetype='text/csv',
                        headers={'Content-Disposition': f'attachment; filename="{filename}"'})
    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


# ── CSV: Profit Report ────────────────────────────────────────────
@reports_bp.route('/api/reports/profit/csv', methods=['GET'])
@roles_required('admin', 'manager')
def profit_report_csv():
    date_from, date_to, err = _parse_dates(request.args)
    if err:
        return jsonify({'error': err}), 400

    conn = _get_conn()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cur = conn.cursor(dictionary=True)

        cur.execute("""
            SELECT bi.product_id,
                   COALESCE(bi.product_name, p.name) AS product_name,
                   SUM(bi.quantity) AS quantity_sold,
                   SUM(bi.item_total) AS revenue,
                   SUM(bi.quantity * p.cost_price) AS cost,
                   SUM(bi.item_total - bi.quantity * p.cost_price) AS profit
            FROM bill_items bi
            JOIN bills b ON b.bill_id = bi.bill_id
            JOIN products p ON p.product_id = bi.product_id
            WHERE b.bill_date BETWEEN %s AND %s
              AND p.cost_price IS NOT NULL
            GROUP BY bi.product_id, COALESCE(bi.product_name, p.name)
            ORDER BY profit DESC
        """, (date_from, date_to))
        rows = cur.fetchall()
        cur.close()

        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(['product_id', 'product_name', 'quantity_sold', 'revenue', 'cost', 'profit', 'margin_percent'])
        for r in rows:
            rev = float(r['revenue'])
            margin = round(float(r['profit']) / rev * 100, 2) if rev > 0 else 0
            writer.writerow([r['product_id'], r['product_name'], r['quantity_sold'],
                             r['revenue'], r['cost'], r['profit'], margin])

        output = buf.getvalue()
        buf.close()
        filename = f"profit_report_{date_from}_to_{date_to}.csv"
        return Response(output, mimetype='text/csv',
                        headers={'Content-Disposition': f'attachment; filename="{filename}"'})
    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


# ── Low Stock Prediction ──────────────────────────────────────────
@reports_bp.route('/api/reports/low-stock-prediction', methods=['GET'])
@roles_required('admin', 'manager')
def low_stock_prediction():
    conn = _get_conn()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cur = conn.cursor(dictionary=True)

        cur.execute("""
            SELECT p.product_id,
                   p.name AS product_name,
                   p.stock_quantity,
                   p.minimum_stock_threshold,
                   COALESCE(rs.total_sold, 0) AS total_sold_30d,
                   ROUND(COALESCE(rs.total_sold, 0) / 30.0, 2) AS avg_daily_sales
            FROM products p
            LEFT JOIN (
                SELECT bi.product_id, SUM(bi.quantity) AS total_sold
                FROM bill_items bi
                JOIN bills b ON b.bill_id = bi.bill_id
                WHERE b.bill_date >= DATE_SUB(CURDATE(), INTERVAL 30 DAY)
                GROUP BY bi.product_id
            ) rs ON rs.product_id = p.product_id
            WHERE p.stock_quantity <= p.minimum_stock_threshold
            ORDER BY p.stock_quantity ASC
        """)
        rows = cur.fetchall()
        cur.close()

        results = []
        for r in rows:
            avg_daily = float(r['avg_daily_sales'])
            if avg_daily > 0:
                est_days = round(float(r['stock_quantity']) / avg_daily, 1)
            else:
                est_days = None
            results.append({
                'product_id': r['product_id'],
                'product_name': r['product_name'],
                'stock_quantity': r['stock_quantity'],
                'minimum_stock_threshold': r['minimum_stock_threshold'],
                'total_sold_30d': r['total_sold_30d'],
                'avg_daily_sales': avg_daily,
                'est_days_remaining': est_days
            })

        return jsonify({'products': results}), 200

    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


# ── Slow Moving Products ──────────────────────────────────────────
@reports_bp.route('/api/reports/slow-moving', methods=['GET'])
@roles_required('admin', 'manager')
def slow_moving_products():
    days = request.args.get('days', 30, type=int)
    if days not in (30, 60, 90):
        return jsonify({'error': 'days must be 30, 60, or 90'}), 400

    conn = _get_conn()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cur = conn.cursor(dictionary=True)

        cur.execute("""
            SELECT p.product_id,
                   p.name AS product_name,
                   p.stock_quantity,
                   MAX(b.bill_date) AS last_sale_date,
                   DATEDIFF(CURDATE(), MAX(b.bill_date)) AS days_since_last_sale
            FROM products p
            LEFT JOIN bill_items bi ON bi.product_id = p.product_id
            LEFT JOIN bills b ON b.bill_id = bi.bill_id
            GROUP BY p.product_id, p.name, p.stock_quantity
            HAVING last_sale_date IS NULL
                OR last_sale_date < DATE_SUB(CURDATE(), INTERVAL %s DAY)
            ORDER BY last_sale_date ASC
        """, (days,))
        rows = cur.fetchall()
        cur.close()

        results = []
        for r in rows:
            results.append({
                'product_id': r['product_id'],
                'product_name': r['product_name'],
                'stock_quantity': r['stock_quantity'],
                'last_sale_date': str(r['last_sale_date']) if r['last_sale_date'] else None,
                'days_since_last_sale': r['days_since_last_sale']
            })

        return jsonify({'days': days, 'products': results}), 200

    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


# ── Dead Stock Report ─────────────────────────────────────────────
@reports_bp.route('/api/reports/dead-stock', methods=['GET'])
@roles_required('admin', 'manager')
def dead_stock_report():
    conn = _get_conn()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cur = conn.cursor(dictionary=True)

        cur.execute("""
            SELECT p.product_id,
                   p.name AS product_name,
                   p.stock_quantity,
                   p.price,
                   MAX(b.bill_date) AS last_sale_date,
                   DATEDIFF(CURDATE(), MAX(b.bill_date)) AS days_since_last_sale,
                   ROUND(p.stock_quantity * p.price, 2) AS inventory_value
            FROM products p
            LEFT JOIN bill_items bi ON bi.product_id = p.product_id
            LEFT JOIN bills b ON b.bill_id = bi.bill_id
            WHERE p.stock_quantity > 0
            GROUP BY p.product_id, p.name, p.stock_quantity, p.price
            HAVING last_sale_date IS NULL
                OR last_sale_date < DATE_SUB(CURDATE(), INTERVAL 90 DAY)
            ORDER BY inventory_value DESC
        """)
        rows = cur.fetchall()
        cur.close()

        results = []
        total_value = 0
        for r in rows:
            val = float(r['inventory_value'])
            total_value += val
            results.append({
                'product_id': r['product_id'],
                'product_name': r['product_name'],
                'stock_quantity': r['stock_quantity'],
                'price': float(r['price']),
                'last_sale_date': str(r['last_sale_date']) if r['last_sale_date'] else None,
                'days_since_last_sale': r['days_since_last_sale'],
                'inventory_value': val
            })

        return jsonify({
            'total_inventory_value': round(total_value, 2),
            'products': results
        }), 200

    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


# ── Purchase Report ────────────────────────────────────────────────
@reports_bp.route('/api/reports/purchases', methods=['GET'])
@roles_required('admin', 'manager')
def purchase_report():
    category_id = request.args.get('category_id')
    if category_id:
        try:
            category_id = int(category_id)
        except (ValueError, TypeError):
            return jsonify({'error': 'category_id must be an integer'}), 400

    supplier_id = request.args.get('supplier_id')
    if supplier_id:
        try:
            supplier_id = int(supplier_id)
        except (ValueError, TypeError):
            return jsonify({'error': 'supplier_id must be an integer'}), 400

    conn = _get_conn()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cur = conn.cursor(dictionary=True)

        cat_filter = ''
        cat_params = ()
        if category_id:
            cat_filter = 'AND poi.product_id IN (SELECT product_id FROM products WHERE category_id = %s)'
            cat_params = (category_id,)

        sup_filter = ''
        sup_params = ()
        if supplier_id:
            sup_filter = 'AND po.supplier_id = %s'
            sup_params = (supplier_id,)

        # Purchases today
        cur.execute(f"""
            SELECT COUNT(*) AS count,
                   COALESCE(SUM(subtotal), 0) AS total_value
            FROM purchase_orders po
            WHERE DATE(po.created_at) = CURDATE()
              AND status IN ('approved','received','partially_received')
              {sup_filter}
        """, sup_params)
        today = cur.fetchone()

        # Purchases this month
        cur.execute(f"""
            SELECT COUNT(*) AS count,
                   COALESCE(SUM(subtotal), 0) AS total_value
            FROM purchase_orders po
            WHERE DATE_FORMAT(po.created_at, '%Y-%m') = DATE_FORMAT(CURDATE(), '%Y-%m')
              AND status IN ('approved','received','partially_received')
              {sup_filter}
        """, sup_params)
        monthly = cur.fetchone()

        # Supplier-wise breakdown (current month)
        cur.execute(f"""
            SELECT po.supplier_id,
                   po.supplier_name,
                   COUNT(*) AS order_count,
                   COALESCE(SUM(po.subtotal), 0) AS total_value
            FROM purchase_orders po
            WHERE DATE_FORMAT(po.created_at, '%Y-%m') = DATE_FORMAT(CURDATE(), '%Y-%m')
              AND po.status IN ('approved','received','partially_received')
              {sup_filter}
            GROUP BY po.supplier_id, po.supplier_name
            ORDER BY total_value DESC
        """, sup_params)
        by_supplier = cur.fetchall()

        # Category-wise breakdown
        cur.execute(f"""
            SELECT COALESCE(c.name, 'Uncategorized') AS category_name,
                   COUNT(DISTINCT poi.po_id) AS order_count,
                   COALESCE(SUM(poi.quantity_ordered), 0) AS total_qty,
                   COALESCE(SUM(poi.item_total), 0) AS total_value
            FROM purchase_order_items poi
            JOIN purchase_orders po ON po.po_id = poi.po_id
            LEFT JOIN products p ON p.product_id = poi.product_id
            LEFT JOIN categories c ON c.category_id = p.category_id
            WHERE DATE_FORMAT(po.created_at, '%Y-%m') = DATE_FORMAT(CURDATE(), '%Y-%m')
              AND po.status IN ('approved','received','partially_received')
              {cat_filter}
              {sup_filter}
            GROUP BY c.name
            ORDER BY total_value DESC
        """, cat_params + sup_params)
        by_category = cur.fetchall()

        # Overall PO stats
        cur.execute(f"""
            SELECT COUNT(*) AS total_pos,
                   COALESCE(SUM(subtotal), 0) AS total_value,
                   SUM(CASE WHEN status IN ('draft','pending_approval','approved') THEN 1 ELSE 0 END) AS pending_count
            FROM purchase_orders po
            WHERE 1=1 {sup_filter}
        """, sup_params)
        overall = cur.fetchone()

        cur.close()
        return jsonify({
            'today': today,
            'monthly': monthly,
            'by_supplier': by_supplier,
            'by_category': by_category,
            'overall': overall
        }), 200

    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


# ── Supplier Report ────────────────────────────────────────────────
@reports_bp.route('/api/reports/suppliers', methods=['GET'])
@roles_required('admin', 'manager')
def supplier_report():
    supplier_id = request.args.get('supplier_id')
    if supplier_id:
        try:
            supplier_id = int(supplier_id)
        except (ValueError, TypeError):
            return jsonify({'error': 'supplier_id must be an integer'}), 400

    conn = _get_conn()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cur = conn.cursor(dictionary=True)

        sup_filter = ''
        sup_params = ()
        if supplier_id:
            sup_filter = 'WHERE s.supplier_id = %s'
            sup_params = (supplier_id,)

        cur.execute(f"""
            SELECT s.supplier_id,
                   s.name AS supplier_name,
                   s.contact_person,
                   s.phone,
                   s.email,
                   COALESCE(po_agg.total_orders, 0) AS total_orders,
                   COALESCE(po_agg.pending_orders, 0) AS pending_orders,
                   COALESCE(po_agg.total_value, 0) AS total_value,
                   po_agg.last_order_date
            FROM suppliers s
            LEFT JOIN (
                SELECT supplier_id,
                       COUNT(po_id) AS total_orders,
                       SUM(CASE WHEN status IN ('draft','pending_approval','approved') THEN 1 ELSE 0 END) AS pending_orders,
                       COALESCE(SUM(CASE WHEN status IN ('received','partially_received') THEN subtotal ELSE 0 END), 0) AS total_value,
                       MAX(created_at) AS last_order_date
                FROM purchase_orders
                GROUP BY supplier_id
            ) po_agg ON po_agg.supplier_id = s.supplier_id
            {sup_filter}
            ORDER BY total_value DESC
        """, sup_params)
        suppliers = cur.fetchall()

        total_suppliers = len(suppliers)
        active_suppliers = sum(1 for s in suppliers if s['total_orders'] > 0)
        total_orders_all = sum(s['total_orders'] for s in suppliers)
        total_value_all = sum(float(s['total_value']) for s in suppliers)

        cur.close()
        return jsonify({
            'summary': {
                'total_suppliers': total_suppliers,
                'active_suppliers': active_suppliers,
                'total_orders': total_orders_all,
                'total_value': round(total_value_all, 2)
            },
            'suppliers': suppliers
        }), 200

    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


# ── KPI Dashboard ─────────────────────────────────────────────────
@reports_bp.route('/api/reports/kpi-dashboard', methods=['GET'])
@roles_required('admin', 'manager', 'staff')
def kpi_dashboard():
    conn = _get_conn()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cur = conn.cursor(dictionary=True)

        # ── Current month bounds ──
        cur.execute("SELECT DATE_FORMAT(CURDATE(), '%Y-%m-01') AS month_start, CURDATE() AS today")
        bounds = cur.fetchone()
        month_start = str(bounds['month_start'])
        today = str(bounds['today'])

        # Previous month bounds
        cur.execute("SELECT DATE_FORMAT(DATE_SUB(CURDATE(), INTERVAL 1 MONTH), '%Y-%m-01') AS prev_start, LAST_DAY(DATE_SUB(CURDATE(), INTERVAL 1 MONTH)) AS prev_end")
        prev = cur.fetchone()
        prev_start = str(prev['prev_start'])
        prev_end = str(prev['prev_end'])

        # ── KPI: Profit (current month, reuses profit report logic) ──
        cur.execute("""
            SELECT COALESCE(SUM(bi.item_total), 0) AS revenue,
                   COALESCE(SUM(bi.quantity * p.cost_price), 0) AS cost,
                   COALESCE(SUM(bi.item_total - bi.quantity * p.cost_price), 0) AS profit
            FROM bill_items bi
            JOIN bills b ON b.bill_id = bi.bill_id
            JOIN products p ON p.product_id = bi.product_id
            WHERE b.bill_date BETWEEN %s AND %s
              AND p.cost_price IS NOT NULL
        """, (month_start, today))
        profit_cur = cur.fetchone()

        # ── KPI: Profit (previous month, for growth calc) ──
        cur.execute("""
            SELECT COALESCE(SUM(bi.item_total - bi.quantity * p.cost_price), 0) AS profit
            FROM bill_items bi
            JOIN bills b ON b.bill_id = bi.bill_id
            JOIN products p ON p.product_id = bi.product_id
            WHERE b.bill_date BETWEEN %s AND %s
              AND p.cost_price IS NOT NULL
        """, (prev_start, prev_end))
        profit_prev = cur.fetchone()

        # ── KPI: Customers (estimated — distinct phone, current month) ──
        cur.execute("""
            SELECT COUNT(DISTINCT customer_phone) AS count
            FROM bills
            WHERE bill_date BETWEEN %s AND %s
              AND customer_phone IS NOT NULL
              AND TRIM(customer_phone) != ''
        """, (month_start, today))
        customers_cur = cur.fetchone()

        # ── KPI: Pending Payments ──
        cur.execute("""
            SELECT COUNT(*) AS count,
                   COALESCE(SUM(amount_due - amount_paid), 0) AS total_outstanding
            FROM pending_payments
            WHERE status IN ('pending', 'partial')
        """)
        pending = cur.fetchone()

        # ── KPI: Revenue (current month + previous month for growth %) ──
        cur.execute("""
            SELECT COALESCE(SUM(grand_total), 0) AS revenue
            FROM bills WHERE bill_date BETWEEN %s AND %s
        """, (month_start, today))
        rev_cur = cur.fetchone()

        cur.execute("""
            SELECT COALESCE(SUM(grand_total), 0) AS revenue
            FROM bills WHERE bill_date BETWEEN %s AND %s
        """, (prev_start, prev_end))
        rev_prev = cur.fetchone()

        # ── Growth % (revenue) ──
        cur_rev = float(rev_cur['revenue'])
        prev_rev = float(rev_prev['revenue'])
        if prev_rev > 0:
            growth_pct = round((cur_rev - prev_rev) / prev_rev * 100, 1)
        else:
            growth_pct = None if cur_rev == 0 else 100.0

        # ── Trend: last 6 months ──
        cur.execute("""
            SELECT DATE_FORMAT(bill_date, '%Y-%m') AS month_label,
                   COUNT(*) AS order_count,
                   COALESCE(SUM(grand_total), 0) AS revenue
            FROM bills
            WHERE bill_date >= DATE_SUB(DATE_FORMAT(CURDATE(), '%Y-%m-01'), INTERVAL 5 MONTH)
            GROUP BY DATE_FORMAT(bill_date, '%Y-%m')
            ORDER BY month_label
        """)
        sales_trend = cur.fetchall()

        # Profit trend (same 6 months)
        cur.execute("""
            SELECT DATE_FORMAT(b.bill_date, '%Y-%m') AS month_label,
                   COALESCE(SUM(bi.item_total), 0) AS revenue,
                   COALESCE(SUM(bi.quantity * p.cost_price), 0) AS cost,
                   COALESCE(SUM(bi.item_total - bi.quantity * p.cost_price), 0) AS profit
            FROM bill_items bi
            JOIN bills b ON b.bill_id = bi.bill_id
            JOIN products p ON p.product_id = bi.product_id
            WHERE b.bill_date >= DATE_SUB(DATE_FORMAT(CURDATE(), '%Y-%m-01'), INTERVAL 5 MONTH)
              AND p.cost_price IS NOT NULL
            GROUP BY DATE_FORMAT(b.bill_date, '%Y-%m')
            ORDER BY month_label
        """)
        profit_trend_raw = cur.fetchall()

        # Build profit trend dict for easy lookup
        profit_map = {r['month_label']: r for r in profit_trend_raw}

        # Merge trends into unified structure
        all_months = sorted(set(r['month_label'] for r in sales_trend) | set(profit_map.keys()))
        trend = []
        for m in all_months:
            s = next((r for r in sales_trend if r['month_label'] == m), None)
            p = profit_map.get(m)
            orders = int(s['order_count']) if s else 0
            rev = float(s['revenue']) if s else 0
            prof = float(p['profit']) if p else 0
            aov = round(rev / orders, 2) if orders > 0 else 0
            trend.append({
                'month': m,
                'orders': orders,
                'revenue': round(rev, 2),
                'profit': round(prof, 2),
                'aov': aov
            })

        cur.close()

        role = request.current_user.get('role')

        kpis = {
            'profit': round(float(profit_cur['profit']), 2),
            'customers_estimated': int(customers_cur['count']),
            'pending_count': int(pending['count']),
            'pending_outstanding': round(float(pending['total_outstanding']), 2),
            'growth_pct': growth_pct,
            'revenue_current': round(cur_rev, 2),
            'revenue_previous': round(prev_rev, 2)
        }

        if role == 'staff':
            kpis = {
                'pending_count': int(pending['count']),
                'pending_outstanding': round(float(pending['total_outstanding']), 2)
            }
            trend = []

        return jsonify({
            'kpis': kpis,
            'trend': trend
        }), 200

    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


# ── Smart Insights ────────────────────────────────────────────────
@reports_bp.route('/api/reports/insights', methods=['GET'])
@roles_required('admin', 'manager')
def smart_insights():
    conn = _get_conn()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cur = conn.cursor(dictionary=True)
        insights = []

        # ── Month bounds ──
        cur.execute("SELECT DATE_FORMAT(CURDATE(), '%Y-%m-01') AS month_start, CURDATE() AS today")
        bounds = cur.fetchone()
        month_start = str(bounds['month_start'])
        today = str(bounds['today'])

        cur.execute("SELECT DATE_FORMAT(DATE_SUB(CURDATE(), INTERVAL 1 MONTH), '%Y-%m-01') AS prev_start, LAST_DAY(DATE_SUB(CURDATE(), INTERVAL 1 MONTH)) AS prev_end")
        prev = cur.fetchone()
        prev_start = str(prev['prev_start'])
        prev_end = str(prev['prev_end'])

        # ── Rule 1: Revenue trend (>= 5% threshold) ──
        cur.execute("""
            SELECT COALESCE(SUM(grand_total), 0) AS revenue
            FROM bills WHERE bill_date BETWEEN %s AND %s
        """, (month_start, today))
        rev_cur = float(cur.fetchone()['revenue'])

        cur.execute("""
            SELECT COALESCE(SUM(grand_total), 0) AS revenue
            FROM bills WHERE bill_date BETWEEN %s AND %s
        """, (prev_start, prev_end))
        rev_prev = float(cur.fetchone()['revenue'])

        if rev_prev > 0:
            change_pct = round((rev_cur - rev_prev) / rev_prev * 100, 1)
            if abs(change_pct) >= 5:
                direction = 'increased' if change_pct > 0 else 'decreased'
                insight_type = 'positive' if change_pct > 0 else 'warning'
                insights.append({
                    'text': f"Revenue {direction} {abs(change_pct)}% this month compared to last month.",
                    'type': insight_type,
                    'category': 'sales'
                })
        elif rev_cur > 0 and rev_prev == 0:
            insights.append({
                'text': "Revenue started this month after zero revenue last month.",
                'type': 'positive',
                'category': 'sales'
            })

        # ── Rule 2: Per-product sales trend (top 5 by revenue, >= 15% change) ──
        cur.execute("""
            SELECT bi.product_id,
                   COALESCE(bi.product_name, p.name) AS product_name,
                   SUM(bi.quantity) AS qty_cur,
                   SUM(bi.item_total) AS rev_cur
            FROM bill_items bi
            JOIN bills b ON b.bill_id = bi.bill_id
            LEFT JOIN products p ON p.product_id = bi.product_id
            WHERE b.bill_date BETWEEN %s AND %s
            GROUP BY bi.product_id, COALESCE(bi.product_name, p.name)
            ORDER BY rev_cur DESC
            LIMIT 5
        """, (month_start, today))
        top5 = cur.fetchall()

        for prod in top5:
            cur.execute("""
                SELECT COALESCE(SUM(bi.quantity), 0) AS qty_prev
                FROM bill_items bi
                JOIN bills b ON b.bill_id = bi.bill_id
                WHERE b.bill_date BETWEEN %s AND %s
                  AND bi.product_id = %s
            """, (prev_start, prev_end, prod['product_id']))
            qty_prev = int(cur.fetchone()['qty_prev'])
            qty_cur = int(prod['qty_cur'])

            if qty_prev > 0:
                pct = round((qty_cur - qty_prev) / qty_prev * 100, 1)
                if abs(pct) >= 15:
                    direction = 'increased' if pct > 0 else 'decreased'
                    insight_type = 'positive' if pct > 0 else 'warning'
                    name = prod['product_name'] or f"Product #{prod['product_id']}"
                    insights.append({
                        'text': f"{name} sales {direction} {abs(pct)}% this month ({qty_cur} vs {qty_prev} units).",
                        'type': insight_type,
                        'category': 'sales'
                    })
            elif qty_cur > 0 and qty_prev == 0:
                name = prod['product_name'] or f"Product #{prod['product_id']}"
                insights.append({
                    'text': f"{name} had new sales this month ({qty_cur} units) after zero last month.",
                    'type': 'positive',
                    'category': 'sales'
                })

        # ── Rule 3: Stock-finishing warnings (<= 5 days, reuses Phase 11B logic) ──
        cur.execute("""
            SELECT p.product_id,
                   p.name AS product_name,
                   p.stock_quantity,
                   COALESCE(rs.total_sold, 0) AS total_sold_30d,
                   ROUND(COALESCE(rs.total_sold, 0) / 30.0, 2) AS avg_daily_sales
            FROM products p
            LEFT JOIN (
                SELECT bi.product_id, SUM(bi.quantity) AS total_sold
                FROM bill_items bi
                JOIN bills b ON b.bill_id = bi.bill_id
                WHERE b.bill_date >= DATE_SUB(CURDATE(), INTERVAL 30 DAY)
                GROUP BY bi.product_id
            ) rs ON rs.product_id = p.product_id
            WHERE p.stock_quantity <= p.minimum_stock_threshold
            ORDER BY p.stock_quantity ASC
        """)
        low_stock_rows = cur.fetchall()

        for r in low_stock_rows:
            avg_daily = float(r['avg_daily_sales'])
            if avg_daily > 0:
                est_days = round(float(r['stock_quantity']) / avg_daily, 1)
                if est_days <= 5:
                    insights.append({
                        'text': f"Stock of {r['product_name']} may finish in {est_days} days (current: {r['stock_quantity']} units).",
                        'type': 'warning',
                        'category': 'stock'
                    })

        # ── Rule 4: Top customer (approximate, current month) ──
        cur.execute("""
            SELECT customer_phone,
                   SUM(grand_total) AS total_spent,
                   COUNT(*) AS bill_count
            FROM bills
            WHERE bill_date BETWEEN %s AND %s
              AND customer_phone IS NOT NULL
              AND TRIM(customer_phone) != ''
            GROUP BY customer_phone
            ORDER BY total_spent DESC
            LIMIT 1
        """, (month_start, today))
        top_cust = cur.fetchone()
        if top_cust:
            spent = float(top_cust['total_spent'])
            insights.append({
                'text': f"Top customer spent Rs. {spent:,.2f} across {top_cust['bill_count']} bill(s) this month.",
                'type': 'neutral',
                'category': 'customers'
            })

        # ── Rule 5: Best/worst month (last 12 months) ──
        cur.execute("""
            SELECT DATE_FORMAT(bill_date, '%Y-%m') AS month_label,
                   DATE_FORMAT(bill_date, '%M %Y') AS month_name,
                   COALESCE(SUM(grand_total), 0) AS revenue
            FROM bills
            WHERE bill_date >= DATE_SUB(DATE_FORMAT(CURDATE(), '%Y-%m-01'), INTERVAL 11 MONTH)
            GROUP BY DATE_FORMAT(bill_date, '%Y-%m'), DATE_FORMAT(bill_date, '%M %Y')
            ORDER BY month_label
        """)
        monthly_rev = cur.fetchall()

        if len(monthly_rev) >= 2:
            best = max(monthly_rev, key=lambda x: float(x['revenue']))
            worst = min(monthly_rev, key=lambda x: float(x['revenue']))
            if float(best['revenue']) > 0:
                insights.append({
                    'text': f"Best month: {best['month_name']} (Rs. {float(best['revenue']):,.2f}).",
                    'type': 'positive',
                    'category': 'sales'
                })
            if float(worst['revenue']) > 0 and worst['month_label'] != best['month_label']:
                insights.append({
                    'text': f"Worst month: {worst['month_name']} (Rs. {float(worst['revenue']):,.2f}).",
                    'type': 'neutral',
                    'category': 'sales'
                })

        cur.close()
        return jsonify({'insights': insights, 'count': len(insights)}), 200

    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


# ════════════════════════════════════════════════════════════════════
#  PDF EXPORT ENDPOINTS
# ════════════════════════════════════════════════════════════════════

# ── PDF: Sales Report ────────────────────────────────────────────
@reports_bp.route('/api/reports/sales/pdf', methods=['GET'])
@roles_required('admin', 'manager')
def sales_report_pdf():
    date_from, date_to, err = _parse_dates(request.args)
    if err:
        return jsonify({'error': err}), 400

    granularity = request.args.get('granularity', 'daily').lower()
    if granularity not in VALID_GRANULARITIES:
        return jsonify({'error': f'granularity must be one of: {", ".join(VALID_GRANULARITIES)}'}), 400

    conn = _get_conn()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cur = conn.cursor(dictionary=True)
        if granularity == 'daily':
            group_expr = 'bill_date'
            label_expr = 'DATE_FORMAT(bill_date, "%Y-%m-%d")'
        elif granularity == 'weekly':
            group_expr = 'YEARWEEK(bill_date, 1)'
            label_expr = 'CONCAT(YEAR(bill_date), "-W", LPAD(WEEK(bill_date, 1), 2, "0"))'
        elif granularity == 'monthly':
            group_expr = 'DATE_FORMAT(bill_date, "%Y-%m")'
            label_expr = 'DATE_FORMAT(bill_date, "%Y-%m")'
        else:
            group_expr = 'YEAR(bill_date)'
            label_expr = 'YEAR(bill_date)'

        cur.execute(f"""
            SELECT {label_expr} AS period_label,
                   COALESCE(SUM(grand_total), 0) AS revenue,
                   COUNT(*) AS bill_count,
                   COALESCE(SUM(discount_amount), 0) AS discount,
                   COALESCE(SUM(gst_amount), 0) AS gst
            FROM bills
            WHERE bill_date BETWEEN %s AND %s
            GROUP BY {group_expr}
            ORDER BY MIN(bill_date)
        """, (date_from, date_to))
        rows = cur.fetchall()
        cur.close()

        headers = ['Period', 'Revenue', 'Bills', 'Discount', 'GST']
        table_rows = [[r['period_label'], float(r['revenue']), int(r['bill_count']),
                        float(r['discount']), float(r['gst'])] for r in rows]
        pdf_bytes = generate_report_pdf('Sales Report', headers, table_rows,
                                         date_from=date_from, date_to=date_to,
                                         money_cols={1, 3, 4}, right_align_cols={1, 2, 3, 4})
        filename = f"sales_report_{date_from}_to_{date_to}.pdf"
        return Response(pdf_bytes, mimetype='application/pdf',
                        headers={'Content-Disposition': f'attachment; filename="{filename}"'})
    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


# ── PDF: Stock Report ────────────────────────────────────────────
@reports_bp.route('/api/reports/stock/pdf', methods=['GET'])
@roles_required('admin', 'manager')
def stock_report_pdf():
    date_from, date_to, err = _parse_dates(request.args)
    if err:
        return jsonify({'error': err}), 400

    category_id = request.args.get('category_id')
    if category_id:
        try:
            category_id = int(category_id)
        except (ValueError, TypeError):
            return jsonify({'error': 'category_id must be an integer'}), 400

    conn = _get_conn()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cur = conn.cursor(dictionary=True)
        _, by_product, _ = _get_stock_report_data(cur, date_from, date_to)
        cur.close()

        if category_id:
            conn2 = _get_conn()
            if conn2:
                cur2 = conn2.cursor(dictionary=True)
                cur2.execute("SELECT product_id FROM products WHERE category_id = %s", (category_id,))
                pids = {r['product_id'] for r in cur2.fetchall()}
                cur2.close()
                conn2.close()
                by_product = [r for r in by_product if r['product_id'] in pids]

        headers = ['Product', 'Opening', 'In', 'Out', 'Closing', 'Reconciled']
        table_rows = [[r['product_name'], int(r['opening_stock']), int(r['stock_in']),
                        int(r['stock_out']), int(r['closing_stock']),
                        'Yes' if r['reconciled'] else 'No'] for r in by_product]
        pdf_bytes = generate_report_pdf('Stock Report', headers, table_rows,
                                         date_from=date_from, date_to=date_to,
                                         right_align_cols={1, 2, 3, 4})
        filename = f"stock_report_{date_from}_to_{date_to}.pdf"
        return Response(pdf_bytes, mimetype='application/pdf',
                        headers={'Content-Disposition': f'attachment; filename="{filename}"'})
    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


# ── PDF: Profit Report ───────────────────────────────────────────
@reports_bp.route('/api/reports/profit/pdf', methods=['GET'])
@roles_required('admin', 'manager')
def profit_report_pdf():
    date_from, date_to, err = _parse_dates(request.args)
    if err:
        return jsonify({'error': err}), 400

    conn = _get_conn()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT COALESCE(bi.product_name, p.name) AS product_name,
                   SUM(bi.quantity) AS quantity_sold,
                   SUM(bi.item_total) AS revenue,
                   SUM(bi.quantity * p.cost_price) AS cost,
                   SUM(bi.item_total - bi.quantity * p.cost_price) AS profit
            FROM bill_items bi
            JOIN bills b ON b.bill_id = bi.bill_id
            JOIN products p ON p.product_id = bi.product_id
            WHERE b.bill_date BETWEEN %s AND %s AND p.cost_price IS NOT NULL
            GROUP BY bi.product_id, COALESCE(bi.product_name, p.name)
            ORDER BY profit DESC
        """, (date_from, date_to))
        rows = cur.fetchall()
        cur.close()

        headers = ['Product', 'Qty Sold', 'Revenue', 'Cost', 'Profit', 'Margin%']
        table_rows = []
        for r in rows:
            rev = float(r['revenue'])
            margin = round(float(r['profit']) / rev * 100, 2) if rev > 0 else 0
            table_rows.append([r['product_name'], int(r['quantity_sold']),
                                rev, float(r['cost']), float(r['profit']), margin])
        pdf_bytes = generate_report_pdf('Profit Report', headers, table_rows,
                                         date_from=date_from, date_to=date_to,
                                         money_cols={2, 3, 4}, right_align_cols={1, 2, 3, 4, 5})
        filename = f"profit_report_{date_from}_to_{date_to}.pdf"
        return Response(pdf_bytes, mimetype='application/pdf',
                        headers={'Content-Disposition': f'attachment; filename="{filename}"'})
    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


# ── PDF: Low Stock Prediction ───────────────────────────────────
@reports_bp.route('/api/reports/low-stock-prediction/pdf', methods=['GET'])
@roles_required('admin', 'manager')
def low_stock_prediction_pdf():
    conn = _get_conn()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT p.name AS product_name, p.stock_quantity,
                   p.minimum_stock_threshold,
                   COALESCE(rs.total_sold, 0) AS total_sold_30d,
                   ROUND(COALESCE(rs.total_sold, 0) / 30.0, 2) AS avg_daily_sales
            FROM products p
            LEFT JOIN (
                SELECT bi.product_id, SUM(bi.quantity) AS total_sold
                FROM bill_items bi JOIN bills b ON b.bill_id = bi.bill_id
                WHERE b.bill_date >= DATE_SUB(CURDATE(), INTERVAL 30 DAY)
                GROUP BY bi.product_id
            ) rs ON rs.product_id = p.product_id
            WHERE p.stock_quantity <= p.minimum_stock_threshold
            ORDER BY p.stock_quantity ASC
        """)
        rows = cur.fetchall()
        cur.close()

        headers = ['Product', 'Current Stock', 'Threshold', 'Avg Daily Sales', 'Est. Days']
        table_rows = []
        for r in rows:
            avg_d = float(r['avg_daily_sales'])
            est = round(float(r['stock_quantity']) / avg_d, 1) if avg_d > 0 else None
            table_rows.append([r['product_name'], int(r['stock_quantity']),
                                int(r['minimum_stock_threshold']), avg_d,
                                str(est) if est is not None else 'N/A'])
        pdf_bytes = generate_report_pdf('Low Stock Prediction', headers, table_rows,
                                         subtitle='Products at risk of stock-out',
                                         right_align_cols={1, 2, 3, 4})
        filename = "low_stock_prediction.pdf"
        return Response(pdf_bytes, mimetype='application/pdf',
                        headers={'Content-Disposition': f'attachment; filename="{filename}"'})
    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


# ── PDF: Slow Moving Products ───────────────────────────────────
@reports_bp.route('/api/reports/slow-moving/pdf', methods=['GET'])
@roles_required('admin', 'manager')
def slow_moving_products_pdf():
    days = request.args.get('days', 30, type=int)
    if days not in (30, 60, 90):
        return jsonify({'error': 'days must be 30, 60, or 90'}), 400

    conn = _get_conn()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT p.name AS product_name, p.stock_quantity,
                   MAX(b.bill_date) AS last_sale_date,
                   DATEDIFF(CURDATE(), MAX(b.bill_date)) AS days_since_last_sale
            FROM products p
            LEFT JOIN bill_items bi ON bi.product_id = p.product_id
            LEFT JOIN bills b ON b.bill_id = bi.bill_id
            GROUP BY p.product_id, p.name, p.stock_quantity
            HAVING last_sale_date IS NULL OR last_sale_date < DATE_SUB(CURDATE(), INTERVAL %s DAY)
            ORDER BY last_sale_date ASC
        """, (days,))
        rows = cur.fetchall()
        cur.close()

        headers = ['Product', 'Last Sale', 'Days Since Sale', 'Stock']
        table_rows = [[r['product_name'],
                        str(r['last_sale_date']) if r['last_sale_date'] else 'Never',
                        r['days_since_last_sale'] if r['days_since_last_sale'] is not None else 'N/A',
                        int(r['stock_quantity'])] for r in rows]
        pdf_bytes = generate_report_pdf(f'Slow Moving Products ({days} days)', headers, table_rows,
                                         subtitle=f'No sale in {days}+ days',
                                         right_align_cols={2, 3})
        filename = f"slow_moving_{days}d.pdf"
        return Response(pdf_bytes, mimetype='application/pdf',
                        headers={'Content-Disposition': f'attachment; filename="{filename}"'})
    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


# ── PDF: Dead Stock Report ──────────────────────────────────────
@reports_bp.route('/api/reports/dead-stock/pdf', methods=['GET'])
@roles_required('admin', 'manager')
def dead_stock_report_pdf():
    conn = _get_conn()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT p.name AS product_name, p.stock_quantity, p.price,
                   MAX(b.bill_date) AS last_sale_date,
                   DATEDIFF(CURDATE(), MAX(b.bill_date)) AS days_since_last_sale,
                   ROUND(p.stock_quantity * p.price, 2) AS inventory_value
            FROM products p
            LEFT JOIN bill_items bi ON bi.product_id = p.product_id
            LEFT JOIN bills b ON b.bill_id = bi.bill_id
            WHERE p.stock_quantity > 0
            GROUP BY p.product_id, p.name, p.stock_quantity, p.price
            HAVING last_sale_date IS NULL OR last_sale_date < DATE_SUB(CURDATE(), INTERVAL 90 DAY)
            ORDER BY inventory_value DESC
        """)
        rows = cur.fetchall()
        cur.close()

        headers = ['Product', 'Days Since Sale', 'Stock', 'Unit Price', 'Inventory Value']
        table_rows = [[r['product_name'],
                        r['days_since_last_sale'] if r['days_since_last_sale'] is not None else 'Never',
                        int(r['stock_quantity']), float(r['price']),
                        float(r['inventory_value'])] for r in rows]
        pdf_bytes = generate_report_pdf('Dead Stock Report', headers, table_rows,
                                         subtitle='Unsold for 90+ days',
                                         money_cols={3, 4}, right_align_cols={1, 2, 3, 4})
        filename = "dead_stock_report.pdf"
        return Response(pdf_bytes, mimetype='application/pdf',
                        headers={'Content-Disposition': f'attachment; filename="{filename}"'})
    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


# ── PDF: Purchase Report ─────────────────────────────────────────
@reports_bp.route('/api/reports/purchases/pdf', methods=['GET'])
@roles_required('admin', 'manager')
def purchase_report_pdf():
    category_id = request.args.get('category_id')
    if category_id:
        try:
            category_id = int(category_id)
        except (ValueError, TypeError):
            return jsonify({'error': 'category_id must be an integer'}), 400

    supplier_id = request.args.get('supplier_id')
    if supplier_id:
        try:
            supplier_id = int(supplier_id)
        except (ValueError, TypeError):
            return jsonify({'error': 'supplier_id must be an integer'}), 400

    conn = _get_conn()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cur = conn.cursor(dictionary=True)

        sup_filter = ''
        sup_params = ()
        if supplier_id:
            sup_filter = 'AND po.supplier_id = %s'
            sup_params = (supplier_id,)

        cur.execute(f"""
            SELECT po.po_id, po.supplier_name, po.subtotal, po.status, po.created_at
            FROM purchase_orders po
            WHERE DATE_FORMAT(po.created_at, '%Y-%m') = DATE_FORMAT(CURDATE(), '%Y-%m')
              AND po.status IN ('approved','received','partially_received')
              {sup_filter}
            ORDER BY po.created_at DESC
        """, sup_params)
        rows = cur.fetchall()
        cur.close()

        if category_id:
            conn2 = _get_conn()
            if conn2:
                cur2 = conn2.cursor(dictionary=True)
                cur2.execute("""
                    SELECT DISTINCT poi.po_id FROM purchase_order_items poi
                    JOIN products p ON p.product_id = poi.product_id
                    WHERE p.category_id = %s
                """, (category_id,))
                po_ids = {r['po_id'] for r in cur2.fetchall()}
                cur2.close()
                conn2.close()
                rows = [r for r in rows if r['po_id'] in po_ids]

        headers = ['PO #', 'Supplier', 'Subtotal', 'Status', 'Date']
        table_rows = [[r['po_id'], r['supplier_name'], float(r['subtotal']),
                        r['status'], str(r['created_at'])[:10]] for r in rows]
        pdf_bytes = generate_report_pdf('Purchase Report', headers, table_rows,
                                         money_cols={2}, right_align_cols={0, 2})
        filename = "purchase_report.pdf"
        return Response(pdf_bytes, mimetype='application/pdf',
                        headers={'Content-Disposition': f'attachment; filename="{filename}"'})
    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


# ── PDF: Supplier Report ─────────────────────────────────────────
@reports_bp.route('/api/reports/suppliers/pdf', methods=['GET'])
@roles_required('admin', 'manager')
def supplier_report_pdf():
    supplier_id = request.args.get('supplier_id')
    if supplier_id:
        try:
            supplier_id = int(supplier_id)
        except (ValueError, TypeError):
            return jsonify({'error': 'supplier_id must be an integer'}), 400

    conn = _get_conn()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cur = conn.cursor(dictionary=True)
        sup_filter = ''
        sup_params = ()
        if supplier_id:
            sup_filter = 'WHERE s.supplier_id = %s'
            sup_params = (supplier_id,)

        cur.execute(f"""
            SELECT s.name AS supplier_name, s.contact_person, s.phone, s.email,
                   COUNT(po.po_id) AS total_orders,
                   COALESCE(SUM(CASE WHEN po.status IN ('received','partially_received') THEN po.subtotal ELSE 0 END), 0) AS total_value,
                   MAX(po.created_at) AS last_order_date
            FROM suppliers s
            LEFT JOIN purchase_orders po ON po.supplier_id = s.supplier_id
            {sup_filter}
            GROUP BY s.supplier_id, s.name, s.contact_person, s.phone, s.email
            ORDER BY total_value DESC
        """, sup_params)
        rows = cur.fetchall()
        cur.close()

        headers = ['Supplier', 'Contact', 'Phone', 'Total Orders', 'Total Value', 'Last Order']
        table_rows = [[r['supplier_name'], r['contact_person'] or '', r['phone'] or '',
                        int(r['total_orders']), float(r['total_value']),
                        str(r['last_order_date'])[:10] if r['last_order_date'] else 'Never'] for r in rows]
        pdf_bytes = generate_report_pdf('Supplier Report', headers, table_rows,
                                         money_cols={4}, right_align_cols={3, 4})
        filename = "supplier_report.pdf"
        return Response(pdf_bytes, mimetype='application/pdf',
                        headers={'Content-Disposition': f'attachment; filename="{filename}"'})
    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


# ════════════════════════════════════════════════════════════════════
#  EXCEL EXPORT ENDPOINTS
# ════════════════════════════════════════════════════════════════════

def _excel_response(rows, headers, sheet_name, filename):
    """Helper: build an Excel workbook and return a Flask Response."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if col_idx == 1 else "left")

    for col_idx, h in enumerate(headers, 1):
        max_len = len(str(h)) + 2
        for row_idx in range(2, len(rows) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)) + 2)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(buf.getvalue(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment; filename="{filename}"'})


# ── Excel: Sales Report ─────────────────────────────────────────
@reports_bp.route('/api/reports/sales/excel', methods=['GET'])
@roles_required('admin', 'manager')
def sales_report_excel():
    date_from, date_to, err = _parse_dates(request.args)
    if err:
        return jsonify({'error': err}), 400

    granularity = request.args.get('granularity', 'daily').lower()
    conn = _get_conn()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cur = conn.cursor(dictionary=True)
        if granularity == 'daily':
            label_expr = 'DATE_FORMAT(bill_date, "%Y-%m-%d")'
            group_expr = 'bill_date'
        elif granularity == 'weekly':
            label_expr = 'CONCAT(YEAR(bill_date), "-W", LPAD(WEEK(bill_date, 1), 2, "0"))'
            group_expr = 'YEARWEEK(bill_date, 1)'
        elif granularity == 'monthly':
            label_expr = 'DATE_FORMAT(bill_date, "%Y-%m")'
            group_expr = 'DATE_FORMAT(bill_date, "%Y-%m")'
        else:
            label_expr = 'YEAR(bill_date)'
            group_expr = 'YEAR(bill_date)'

        cur.execute(f"""
            SELECT {label_expr} AS period_label,
                   COALESCE(SUM(grand_total), 0) AS revenue,
                   COUNT(*) AS bill_count,
                   COALESCE(SUM(discount_amount), 0) AS discount,
                   COALESCE(SUM(gst_amount), 0) AS gst
            FROM bills WHERE bill_date BETWEEN %s AND %s
            GROUP BY {group_expr} ORDER BY MIN(bill_date)
        """, (date_from, date_to))
        rows = cur.fetchall()
        cur.close()

        headers = ['Period', 'Revenue', 'Bill Count', 'Discount', 'GST']
        data = [[r['period_label'], float(r['revenue']), int(r['bill_count']),
                  float(r['discount']), float(r['gst'])] for r in rows]
        return _excel_response(data, headers, 'Sales',
                               f"sales_report_{date_from}_to_{date_to}.xlsx")
    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


# ── Excel: Stock Report ─────────────────────────────────────────
@reports_bp.route('/api/reports/stock/excel', methods=['GET'])
@roles_required('admin', 'manager')
def stock_report_excel():
    date_from, date_to, err = _parse_dates(request.args)
    if err:
        return jsonify({'error': err}), 400

    category_id = request.args.get('category_id')
    if category_id:
        try:
            category_id = int(category_id)
        except (ValueError, TypeError):
            return jsonify({'error': 'category_id must be an integer'}), 400

    conn = _get_conn()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cur = conn.cursor(dictionary=True)
        _, by_product, _ = _get_stock_report_data(cur, date_from, date_to)
        cur.close()

        if category_id:
            conn2 = _get_conn()
            if conn2:
                cur2 = conn2.cursor(dictionary=True)
                cur2.execute("SELECT product_id FROM products WHERE category_id = %s", (category_id,))
                pids = {r['product_id'] for r in cur2.fetchall()}
                cur2.close()
                conn2.close()
                by_product = [r for r in by_product if r['product_id'] in pids]

        headers = ['Product', 'Opening', 'In', 'Out', 'Closing', 'Reconciled']
        data = [[r['product_name'], int(r['opening_stock']), int(r['stock_in']),
                  int(r['stock_out']), int(r['closing_stock']),
                  'Yes' if r['reconciled'] else 'No'] for r in by_product]
        return _excel_response(data, headers, 'Stock',
                               f"stock_report_{date_from}_to_{date_to}.xlsx")
    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


# ── Excel: Profit Report ────────────────────────────────────────
@reports_bp.route('/api/reports/profit/excel', methods=['GET'])
@roles_required('admin', 'manager')
def profit_report_excel():
    date_from, date_to, err = _parse_dates(request.args)
    if err:
        return jsonify({'error': err}), 400

    conn = _get_conn()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT COALESCE(bi.product_name, p.name) AS product_name,
                   SUM(bi.quantity) AS quantity_sold,
                   SUM(bi.item_total) AS revenue,
                   SUM(bi.quantity * p.cost_price) AS cost,
                   SUM(bi.item_total - bi.quantity * p.cost_price) AS profit
            FROM bill_items bi
            JOIN bills b ON b.bill_id = bi.bill_id
            JOIN products p ON p.product_id = bi.product_id
            WHERE b.bill_date BETWEEN %s AND %s AND p.cost_price IS NOT NULL
            GROUP BY bi.product_id, COALESCE(bi.product_name, p.name)
            ORDER BY profit DESC
        """, (date_from, date_to))
        rows = cur.fetchall()
        cur.close()

        headers = ['Product', 'Qty Sold', 'Revenue', 'Cost', 'Profit', 'Margin%']
        data = []
        for r in rows:
            rev = float(r['revenue'])
            margin = round(float(r['profit']) / rev * 100, 2) if rev > 0 else 0
            data.append([r['product_name'], int(r['quantity_sold']),
                          rev, float(r['cost']), float(r['profit']), margin])
        return _excel_response(data, headers, 'Profit',
                               f"profit_report_{date_from}_to_{date_to}.xlsx")
    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


# ── Excel: Low Stock Prediction ─────────────────────────────────
@reports_bp.route('/api/reports/low-stock-prediction/excel', methods=['GET'])
@roles_required('admin', 'manager')
def low_stock_prediction_excel():
    conn = _get_conn()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT p.name AS product_name, p.stock_quantity,
                   p.minimum_stock_threshold,
                   ROUND(COALESCE(rs.total_sold, 0) / 30.0, 2) AS avg_daily_sales
            FROM products p
            LEFT JOIN (
                SELECT bi.product_id, SUM(bi.quantity) AS total_sold
                FROM bill_items bi JOIN bills b ON b.bill_id = bi.bill_id
                WHERE b.bill_date >= DATE_SUB(CURDATE(), INTERVAL 30 DAY)
                GROUP BY bi.product_id
            ) rs ON rs.product_id = p.product_id
            WHERE p.stock_quantity <= p.minimum_stock_threshold
            ORDER BY p.stock_quantity ASC
        """)
        rows = cur.fetchall()
        cur.close()

        headers = ['Product', 'Current Stock', 'Threshold', 'Avg Daily Sales', 'Est. Days']
        data = []
        for r in rows:
            avg_d = float(r['avg_daily_sales'])
            est = round(float(r['stock_quantity']) / avg_d, 1) if avg_d > 0 else None
            data.append([r['product_name'], int(r['stock_quantity']),
                          int(r['minimum_stock_threshold']), avg_d,
                          est if est is not None else 'N/A'])
        return _excel_response(data, headers, 'Low Stock', "low_stock_prediction.xlsx")
    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


# ── Excel: Slow Moving Products ─────────────────────────────────
@reports_bp.route('/api/reports/slow-moving/excel', methods=['GET'])
@roles_required('admin', 'manager')
def slow_moving_products_excel():
    days = request.args.get('days', 30, type=int)
    if days not in (30, 60, 90):
        return jsonify({'error': 'days must be 30, 60, or 90'}), 400

    conn = _get_conn()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT p.name AS product_name, p.stock_quantity,
                   MAX(b.bill_date) AS last_sale_date,
                   DATEDIFF(CURDATE(), MAX(b.bill_date)) AS days_since_last_sale
            FROM products p
            LEFT JOIN bill_items bi ON bi.product_id = p.product_id
            LEFT JOIN bills b ON b.bill_id = bi.bill_id
            GROUP BY p.product_id, p.name, p.stock_quantity
            HAVING last_sale_date IS NULL OR last_sale_date < DATE_SUB(CURDATE(), INTERVAL %s DAY)
            ORDER BY last_sale_date ASC
        """, (days,))
        rows = cur.fetchall()
        cur.close()

        headers = ['Product', 'Last Sale', 'Days Since Sale', 'Stock']
        data = [[r['product_name'],
                  str(r['last_sale_date'])[:10] if r['last_sale_date'] else 'Never',
                  r['days_since_last_sale'] if r['days_since_last_sale'] is not None else 'N/A',
                  int(r['stock_quantity'])] for r in rows]
        return _excel_response(data, headers, 'Slow Moving',
                               f"slow_moving_{days}d.xlsx")
    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


# ── Excel: Dead Stock Report ────────────────────────────────────
@reports_bp.route('/api/reports/dead-stock/excel', methods=['GET'])
@roles_required('admin', 'manager')
def dead_stock_report_excel():
    conn = _get_conn()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT p.name AS product_name, p.stock_quantity, p.price,
                   DATEDIFF(CURDATE(), MAX(b.bill_date)) AS days_since_last_sale,
                   ROUND(p.stock_quantity * p.price, 2) AS inventory_value
            FROM products p
            LEFT JOIN bill_items bi ON bi.product_id = p.product_id
            LEFT JOIN bills b ON b.bill_id = bi.bill_id
            WHERE p.stock_quantity > 0
            GROUP BY p.product_id, p.name, p.stock_quantity, p.price
            HAVING MAX(b.bill_date) IS NULL OR MAX(b.bill_date) < DATE_SUB(CURDATE(), INTERVAL 90 DAY)
            ORDER BY inventory_value DESC
        """)
        rows = cur.fetchall()
        cur.close()

        headers = ['Product', 'Days Since Sale', 'Stock', 'Unit Price', 'Inventory Value']
        data = [[r['product_name'],
                  r['days_since_last_sale'] if r['days_since_last_sale'] is not None else 'Never',
                  int(r['stock_quantity']), float(r['price']),
                  float(r['inventory_value'])] for r in rows]
        return _excel_response(data, headers, 'Dead Stock', "dead_stock_report.xlsx")
    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


# ── Excel: Purchase Report ──────────────────────────────────────
@reports_bp.route('/api/reports/purchases/excel', methods=['GET'])
@roles_required('admin', 'manager')
def purchase_report_excel():
    category_id = request.args.get('category_id')
    if category_id:
        try:
            category_id = int(category_id)
        except (ValueError, TypeError):
            return jsonify({'error': 'category_id must be an integer'}), 400

    supplier_id = request.args.get('supplier_id')
    if supplier_id:
        try:
            supplier_id = int(supplier_id)
        except (ValueError, TypeError):
            return jsonify({'error': 'supplier_id must be an integer'}), 400

    conn = _get_conn()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cur = conn.cursor(dictionary=True)
        sup_filter = ''
        sup_params = ()
        if supplier_id:
            sup_filter = 'AND po.supplier_id = %s'
            sup_params = (supplier_id,)

        cur.execute(f"""
            SELECT po.po_id, po.supplier_name, po.subtotal, po.status, po.created_at
            FROM purchase_orders po
            WHERE DATE_FORMAT(po.created_at, '%Y-%m') = DATE_FORMAT(CURDATE(), '%Y-%m')
              AND po.status IN ('approved','received','partially_received')
              {sup_filter}
            ORDER BY po.created_at DESC
        """, sup_params)
        rows = cur.fetchall()
        cur.close()

        if category_id:
            conn2 = _get_conn()
            if conn2:
                cur2 = conn2.cursor(dictionary=True)
                cur2.execute("""
                    SELECT DISTINCT poi.po_id FROM purchase_order_items poi
                    JOIN products p ON p.product_id = poi.product_id
                    WHERE p.category_id = %s
                """, (category_id,))
                po_ids = {r['po_id'] for r in cur2.fetchall()}
                cur2.close()
                conn2.close()
                rows = [r for r in rows if r['po_id'] in po_ids]

        headers = ['PO #', 'Supplier', 'Subtotal', 'Status', 'Date']
        data = [[r['po_id'], r['supplier_name'], float(r['subtotal']),
                  r['status'], str(r['created_at'])[:10]] for r in rows]
        return _excel_response(data, headers, 'Purchases', "purchase_report.xlsx")
    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


# ── Excel: Supplier Report ──────────────────────────────────────
@reports_bp.route('/api/reports/suppliers/excel', methods=['GET'])
@roles_required('admin', 'manager')
def supplier_report_excel():
    supplier_id = request.args.get('supplier_id')
    if supplier_id:
        try:
            supplier_id = int(supplier_id)
        except (ValueError, TypeError):
            return jsonify({'error': 'supplier_id must be an integer'}), 400

    conn = _get_conn()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cur = conn.cursor(dictionary=True)
        sup_filter = ''
        sup_params = ()
        if supplier_id:
            sup_filter = 'WHERE s.supplier_id = %s'
            sup_params = (supplier_id,)

        cur.execute(f"""
            SELECT s.name AS supplier_name, s.contact_person, s.phone, s.email,
                   COUNT(po.po_id) AS total_orders,
                   COALESCE(SUM(CASE WHEN po.status IN ('received','partially_received') THEN po.subtotal ELSE 0 END), 0) AS total_value,
                   MAX(po.created_at) AS last_order_date
            FROM suppliers s
            LEFT JOIN purchase_orders po ON po.supplier_id = s.supplier_id
            {sup_filter}
            GROUP BY s.supplier_id, s.name, s.contact_person, s.phone, s.email
            ORDER BY total_value DESC
        """, sup_params)
        rows = cur.fetchall()
        cur.close()

        headers = ['Supplier', 'Contact', 'Phone', 'Email', 'Total Orders', 'Total Value', 'Last Order']
        data = [[r['supplier_name'], r['contact_person'] or '', r['phone'] or '',
                  r['email'] or '', int(r['total_orders']), float(r['total_value']),
                  str(r['last_order_date'])[:10] if r['last_order_date'] else 'Never'] for r in rows]
        return _excel_response(data, headers, 'Suppliers', "supplier_report.xlsx")
    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()
